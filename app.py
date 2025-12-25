from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session
from flask_cors import CORS
import cv2
import numpy as np
import os
import json
from datetime import datetime, time
import pytz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)
app.secret_key = "super-secret-key-123"

# IST timezone
ist = pytz.timezone('Asia/Kolkata')

# Global vars - IMPROVED CASCADE LOADING
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
recognizer = cv2.face.LBPHFaceRecognizer_create()

# Create folders and JSON files
os.makedirs('training', exist_ok=True)
os.makedirs('data', exist_ok=True)


def preprocess_face(gray_face, size=(200, 200)):
    """Normalize illumination and noise for stable recognition."""
    # Histogram equalization smooths lighting differences
    equalized = cv2.equalizeHist(gray_face)
    # Bilateral filter denoises while preserving edges
    denoised = cv2.bilateralFilter(equalized, 5, 75, 75)
    return cv2.resize(denoised, size)

ADMIN_EMAIL = "admin@company.com"
ADMIN_PASSWORD = "admin123"

# Initialize JSON files
def init_json_files():
    if not os.path.exists('data/students.json'):
        with open('data/students.json', 'w') as f:
            json.dump([], f)
    if not os.path.exists('data/attendance.json'):
        with open('data/attendance.json', 'w') as f:
            json.dump([], f)

init_json_files()

def load_students():
    with open('data/students.json', 'r') as f:
        return json.load(f)

def save_students(students):
    with open('data/students.json', 'w') as f:
        json.dump(students, f, indent=2)

def load_attendance():
    with open('data/attendance.json', 'r') as f:
        return json.load(f)

def save_attendance(attendance):
    with open('data/attendance.json', 'w') as f:
        json.dump(attendance, f, indent=2)

def is_logged_in():
    return session.get("logged_in") is True

@app.route('/')
def index():
    if not is_logged_in():
        return redirect(url_for('login'))
    return render_template('dashboard.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        if email == ADMIN_EMAIL and password == ADMIN_PASSWORD:
            session['logged_in'] = True
            session['admin_email'] = email
            return redirect(url_for('index'))
        else:
            error = "Invalid email or password"
            return render_template('login.html', error=error)
    return render_template('login.html', error=None)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/studentslist')
def students_list():
    students = load_students()
    return jsonify([{'id': s['rollno'], 'name': s['name']} for s in students])

@app.route('/api/registerstudent', methods=['POST'])
def register_student():
    try:
        data = request.json
        rollno = data.get('studentid', '').strip().upper()
        name = data.get('studentname', '').strip()
        
        if not rollno or not name:
            return jsonify({'success': False, 'message': 'Please enter both roll number and name!'}), 400
        
        students = load_students()
        if any(s['rollno'] == rollno for s in students):
            return jsonify({'success': False, 'message': f'Student {rollno} already registered'}), 400
        
        students.append({
            'rollno': rollno,
            'name': name,
            'registered': datetime.now(ist).isoformat()
        })
        save_students(students)
        
        # --- IMPROVED CAMERA CAPTURE (60 images like working project) ---
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            return jsonify({'success': False, 'message': 'Could not open camera'}), 500
        
        student_dir = f'training/{rollno}_{name.replace(" ", "_")}'
        os.makedirs(student_dir, exist_ok=True)
        
        print(f"Capturing 60 images for {name} ({rollno})...")
        print("Look at the camera and move slightly for better training data")
        
        count = 0
        target_images = 60  # INCREASED FROM 5 TO 60
        
        while count < target_images:
            ret, frame = cap.read()
            if not ret:
                break
            
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            
            # IMPROVED DETECTION PARAMETERS
            faces = face_cascade.detectMultiScale(
                gray, 
                scaleFactor=1.2,  # Better than 1.3
                minNeighbors=5,    # More strict
                minSize=(30, 30)   # Minimum face size
            )
            
            for (x, y, w, h) in faces:
                # Extract and normalize face
                face = gray[y:y+h, x:x+w]
                face_resized = preprocess_face(face)
                
                # Draw rectangle
                cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)
                
                # AUTOMATIC CAPTURE (no need to press space)
                cv2.imwrite(f'{student_dir}/{count}.jpg', face_resized)
                count += 1
                
                # Display progress
                progress_text = f"Capturing: {count}/{target_images}"
                cv2.putText(frame, progress_text, (10, 30),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                
                if count >= target_images:
                    break
            
            cv2.imshow('Registering Student - Press ESC to cancel', frame)
            
            key = cv2.waitKey(100) & 0xFF  # Slight delay for better capture
            if key == 27:  # ESC to cancel
                cap.release()
                cv2.destroyAllWindows()
                return jsonify({'success': False, 'message': 'Registration cancelled'}), 400
        
        cap.release()
        cv2.destroyAllWindows()
        
        if count < 30:  # Minimum 30 images required
            return jsonify({
                'success': False, 
                'message': f'Only captured {count} images. Need at least 30. Please try again with better lighting.'
            }), 400
        
        return jsonify({
            'success': True,
            'message': f'Registered {name} ({rollno}). Captured {count} images. Click Train Model next.'
        })
        
    except Exception as e:
        print('Error in register_student:', e)
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

def augment_face(face_img):
    """Apply slight variations to increase training diversity."""
    augmented = []
    
    # Original
    augmented.append(face_img)
    
    # Slight brightness variations
    bright = cv2.convertScaleAbs(face_img, alpha=1.1, beta=10)
    augmented.append(bright)
    
    dark = cv2.convertScaleAbs(face_img, alpha=0.9, beta=-10)
    augmented.append(dark)
    
    # Slight contrast
    contrast = cv2.convertScaleAbs(face_img, alpha=1.05, beta=0)
    augmented.append(contrast)
    
    return augmented

@app.route('/api/trainmodel', methods=['POST'])
def train_model():
    try:
        students = load_students()
        if not students:
            return jsonify({'success': False, 'message': 'No students registered yet'}), 400
        
        # Collect all student data first
        student_data = []  # List of (label_id, rollno, name, image_paths)
        label_id = 0
        
        print("="*60)
        print("TRAINING MODEL - BALANCED & IMPROVED METHOD")
        print("="*60)
        
        # First pass: collect all image paths per student
        for student in students:
            rollno = student['rollno']
            candidate_dirs = [
                f'training/{rollno}_{student["name"].replace(" ", "_")}',
                f'training/{rollno}'
            ]
            image_paths = []
            for d in candidate_dirs:
                if os.path.exists(d):
                    image_paths.extend(
                        [os.path.join(d, f) for f in os.listdir(d) if f.lower().endswith('.jpg')]
                    )
            
            if image_paths:
                student_data.append((label_id, rollno, student['name'], sorted(image_paths)))
                print(f"✓ Found {len(image_paths)} images for {student['name']} (will be Label {label_id})")
                label_id += 1
        
        if not student_data:
            return jsonify({'success': False, 'message': 'No training images found'}), 400
        
        # Find minimum images per student for balanced training
        min_images = min(len(paths) for _, _, _, paths in student_data)
        print(f"\nMinimum images per student: {min_images}")
        print("Using balanced sampling to prevent bias...")
        
        # Second pass: load images with balanced sampling
        faces = []
        labels = []
        label_map = {}  # Will use STRING keys from the start
        
        # Use same number of images per student (prevents bias)
        samples_per_student = min(min_images, 50)  # Use up to 50 images per student
        
        for label_id, rollno, name, image_paths in student_data:
            # IMPORTANT: Use string keys to match JSON serialization
            label_map[str(label_id)] = rollno
            
            # Sample evenly across all images (not just first N)
            if len(image_paths) > samples_per_student:
                # Take evenly spaced samples
                step = len(image_paths) / samples_per_student
                sampled_paths = [image_paths[int(i * step)] for i in range(samples_per_student)]
            else:
                sampled_paths = image_paths
            
            images_loaded = 0
            for path in sampled_paths:
                face_img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
                if face_img is None:
                    continue
                
                # Detect face in image
                detected_faces = face_cascade.detectMultiScale(
                    face_img,
                    scaleFactor=1.2,
                    minNeighbors=5,
                    minSize=(30, 30)
                )
                
                if len(detected_faces) > 0:
                    # Use detected face region
                    for (x, y, w, h) in detected_faces:
                        face_region = face_img[y:y+h, x:x+w]
                        face_processed = preprocess_face(face_region)
                        
                        # Add original + augmented versions
                        augmented_faces = augment_face(face_processed)
                        for aug_face in augmented_faces:
                            faces.append(aug_face)
                            labels.append(label_id)
                            images_loaded += 1
                else:
                    # No face detected, use whole image
                    face_processed = preprocess_face(face_img)
                    augmented_faces = augment_face(face_processed)
                    for aug_face in augmented_faces:
                        faces.append(aug_face)
                        labels.append(label_id)
                        images_loaded += 1
            
            print(f"✓ Loaded {images_loaded} training samples for {name} (Label {label_id})")
        
        if len(faces) == 0:
            return jsonify({'success': False, 'message': 'No valid training images found'}), 400
        
        # Count samples per label to verify balance
        label_counts = {}
        for lbl in labels:
            label_counts[lbl] = label_counts.get(lbl, 0) + 1
        
        print(f"\nSamples per student:")
        for label_id, rollno, name, _ in student_data:
            count = label_counts.get(label_id, 0)
            print(f"  {name} (Label {label_id}): {count} samples")
        
        # Check if balanced (within 10% difference)
        counts = list(label_counts.values())
        if len(counts) > 1:
            min_count = min(counts)
            max_count = max(counts)
            if max_count > min_count * 1.1:  # More than 10% difference
                print(f"\n⚠ WARNING: Data imbalance detected! {max_count} vs {min_count} samples")
                print("  This may cause recognition bias. Consider balancing the data.")
        
        # Shuffle data to prevent order bias
        indices = np.random.permutation(len(faces))
        faces_shuffled = [faces[i] for i in indices]
        labels_shuffled = [labels[i] for i in indices]
        
        print(f"\nTraining with {len(faces_shuffled)} face samples from {len(label_map)} students...")
        print(f"Average samples per student: {len(faces_shuffled) // len(label_map)}")
        
        # Create new recognizer with optimized parameters
        recognizer_new = cv2.face.LBPHFaceRecognizer_create(
            radius=1,
            neighbors=8,
            grid_x=8,
            grid_y=8,
            threshold=70.0  # Match the recognition threshold
        )
        
        # TRAIN WITH SHUFFLED DATA
        recognizer_new.train(faces_shuffled, np.array(labels_shuffled))
        recognizer_new.save('trainer.yml')
        
        # Update global recognizer
        global recognizer
        recognizer = recognizer_new
        
        # Save label map with STRING keys
        with open('label_map.json', 'w') as f:
            json.dump(label_map, f, indent=2)
        
        print("="*60)
        print("TRAINING COMPLETED SUCCESSFULLY")
        print(f"Label map: {label_map}")
        
        # VERIFICATION: Test the model on a few sample images
        print("\n" + "="*60)
        print("VERIFICATION: Testing model on sample images...")
        print("="*60)
        
        verification_results = {}
        for label_id, rollno, name, image_paths in student_data:
            # Test on 3 random images from each student
            test_paths = np.random.choice(image_paths, min(3, len(image_paths)), replace=False)
            correct = 0
            total = 0
            
            for path in test_paths:
                face_img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
                if face_img is None:
                    continue
                
                face_processed = preprocess_face(face_img)
                predicted_label, confidence = recognizer_new.predict(face_processed)
                total += 1
                
                if predicted_label == label_id:
                    correct += 1
                    print(f"  ✓ {name}: Predicted={predicted_label} (confidence={confidence:.1f}) - CORRECT")
                else:
                    predicted_rollno = label_map.get(str(predicted_label), '?')
                    predicted_student = find_student_by_roll(predicted_rollno)
                    predicted_name = predicted_student['name'] if predicted_student else 'Unknown'
                    print(f"  ✗ {name}: Predicted={predicted_label} ({predicted_name}, confidence={confidence:.1f}) - WRONG!")
            
            accuracy = (correct / total * 100) if total > 0 else 0
            verification_results[name] = f"{correct}/{total} ({accuracy:.0f}%)"
            print(f"  {name} accuracy: {correct}/{total} ({accuracy:.0f}%)")
        
        print("="*60)
        
        # Warning if verification shows issues
        for name, result in verification_results.items():
            if "0%" in result or "33%" in result:
                print(f"\n⚠ WARNING: {name} has low verification accuracy!")
                print("  The model may not recognize this student correctly.")
                print("  Consider re-registering this student with better lighting.")
        
        return jsonify({
            'success': True,
            'message': f'Model trained successfully with {len(faces_shuffled)} samples from {len(label_map)} students. Verification: {verification_results}'
        })
        
    except Exception as e:
        print('Training error:', e)
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Training error: {str(e)}'}), 500


# ---------- ATTENDANCE HELPERS ----------
def get_today_date_str():
    return datetime.now(ist).strftime('%Y-%m-%d')

def get_now_time_str():
    return datetime.now(ist).strftime('%H:%M:%S')

def get_now_time_obj():
    return datetime.now(ist).time()

def find_student_by_roll(rollno):
    students = load_students()
    for s in students:
        if s['rollno'] == rollno:
            return s
    return None

def get_today_records():
    today = get_today_date_str()
    all_att = load_attendance()
    return [r for r in all_att if r['date'] == today]

def save_or_update_attendance(rollno, status, intime=None, outtime=None):
    all_att = load_attendance()
    today = get_today_date_str()
    
    # find existing record for this student today
    rec = None
    for r in all_att:
        if r['date'] == today and r['rollno'] == rollno:
            rec = r
            break
    student = find_student_by_roll(rollno)

    # if no record yet, create one
    if rec is None:
        rec = {
            'date': today,
            'rollno': rollno,
            'name': student['name'] if student else '',
            'intime': intime or get_now_time_str(),
            'outtime': outtime or '',
            'status': status
        }
        all_att.append(rec)
    else:
        # update existing
        if intime is not None:
            rec['intime'] = intime
        if outtime is not None:
            rec['outtime'] = outtime
        rec['status'] = status
        if 'name' not in rec:
            rec['name'] = student['name'] if student else ''

    save_attendance(all_att)
    return rec

def decide_attendance_action(existing_rec):
    """
    Returns (action, status) where:
    action: 'IN', 'OUT', 'CLOSED'
    status: 'PRESENT' / 'HALF DAY' / None
    """
    now_t = get_now_time_obj()
    
    # Time windows
    in_start = time(7, 45)
    in_end = time(12, 0)
    out_open = time(12, 0)
    half_start = time(12, 1)
    half_end = time(14, 0)
    full_start = time(14, 50)
    full_end = time(17, 0)
    close_time = time(17, 0)
    
    # 1) No record yet -> only IN allowed in window
    if existing_rec is None:
        if not (in_start <= now_t <= in_end):
            return 'CLOSED', None
        return 'IN', 'PRESENT'
    
    # 2) Has IN but no OUT yet
    if existing_rec.get('intime') and not existing_rec.get('outtime'):
        if now_t > close_time:
            return 'CLOSED', None
        if now_t <= out_open:
            return 'CLOSED', None
        if half_start <= now_t <= half_end:
            return 'OUT', 'HALF DAY'
        if full_start <= now_t <= full_end:
            return 'OUT', 'PRESENT'
        return 'CLOSED', None
    
    # 3) Already has IN and OUT
    return 'CLOSED', existing_rec.get('status')
@app.route('/api/takeattendance', methods=['POST'])
def take_attendance():
    try:
        # Load trained model and label map
        if not os.path.exists('trainer.yml') or not os.path.exists('label_map.json'):
            return jsonify({'success': False, 'message': 'Model not trained yet'}), 400
        
        recognizer.read('trainer.yml')
        with open('label_map.json', 'r') as f:
            label_map = json.load(f)
        
        students = load_students()
        if not students:
            return jsonify({'success': False, 'message': 'No students registered'}), 400
        
        # Open camera
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            return jsonify({'success': False, 'message': 'Could not open camera'}), 500
        
        recognized_rollno = None
        recognized_name = None
        confidence_val = None
        
        print("="*60)
        print("FACE RECOGNITION ATTENDANCE - IMPROVED")
        print("="*60)
        print("Face will be recognized automatically")
        print("Press SPACE to confirm | Press ESC to cancel")
        print("-"*60)
        
        frame_count = 0
        # IMPROVED: Faster recognition with tighter threshold
        recognition_votes = {}  # {rollno: [confidence1, confidence2, ...]}
        
        # IMPROVED THRESHOLDS
        CONFIDENCE_THRESHOLD = 70  # Stricter threshold (lower = better match in LBPH)
        VOTES_REQUIRED = 5  # Reduced from 15 to 5 for faster recognition
        FRAMES_TO_CLEAR = 5  # Clear old votes faster (was 15)
        
        # Track frames since last good recognition
        frames_since_recognition = 0
        
        # Debug: Print label map
        print(f"Label map loaded: {label_map}")
        print(f"Total students in map: {len(label_map)}")
        print(f"Confidence threshold: < {CONFIDENCE_THRESHOLD} (lower is better)")
        print(f"Votes required: {VOTES_REQUIRED}")
        
        while True:
            ret, frame = cap.read()
            if not ret:
                break
            
            frame_count += 1
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            
            # Detect faces
            faces = face_cascade.detectMultiScale(
                gray, 
                scaleFactor=1.2,
                minNeighbors=5,
                minSize=(30, 30)
            )
            
            # Current frame recognition result
            current_rollno = None
            current_name = None
            current_confidence = None
            
            # Process each face
            for (x, y, w, h) in faces:
                face = gray[y:y+h, x:x+w]
                face_resized = preprocess_face(face)
                
                # PREDICT
                label, confidence = recognizer.predict(face_resized)
                
                # Get student info - try both string and int keys
                rollno = None
                if str(label) in label_map:
                    rollno = label_map[str(label)]
                elif label in label_map:
                    rollno = label_map[label]
                
                student = find_student_by_roll(rollno) if rollno is not None else None
                
                # Debug output every 30 frames
                if frame_count % 30 == 0:
                    print(f"Frame {frame_count}: Label={label}, Confidence={confidence:.1f}, Rollno={rollno}, Student={student['name'] if student else 'None'}")
                
                # Check if this is a good match
                if student and confidence < CONFIDENCE_THRESHOLD:
                    # GOOD MATCH - add vote
                    if rollno not in recognition_votes:
                        recognition_votes[rollno] = []
                    recognition_votes[rollno].append(confidence)
                    
                    # Keep only recent votes (last 20 frames)
                    if len(recognition_votes[rollno]) > 20:
                        recognition_votes[rollno].pop(0)
                    
                    frames_since_recognition = 0
                    current_rollno = rollno
                    current_name = student['name']
                    current_confidence = confidence
                else:
                    # NOT A GOOD MATCH - increment counter
                    frames_since_recognition += 1
                    
                    # CLEAR OLD VOTES if we haven't had a good match recently
                    if frames_since_recognition >= FRAMES_TO_CLEAR:
                        recognition_votes.clear()
                        frames_since_recognition = 0
            
            # Determine who to display
            display_rollno = None
            display_name = None
            display_status = "Unknown"
            display_color = (0, 0, 255)  # Red for unknown
            vote_count = 0
            avg_confidence = 0
            
            if recognition_votes:
                # Find the student with most votes
                best_rollno = max(recognition_votes.keys(), key=lambda k: len(recognition_votes[k]))
                vote_count = len(recognition_votes[best_rollno])
                avg_confidence = sum(recognition_votes[best_rollno]) / vote_count
                
                # Check if we have enough votes
                if vote_count >= VOTES_REQUIRED:
                    student = find_student_by_roll(best_rollno)
                    display_rollno = best_rollno
                    display_name = student['name']
                    display_status = "READY - Press SPACE"
                    display_color = (0, 255, 0)  # Green
                    
                    # Auto-set for confirmation
                    recognized_rollno = best_rollno
                    recognized_name = student['name']
                    confidence_val = 100 - avg_confidence
                elif vote_count > 0:
                    # Still analyzing
                    student = find_student_by_roll(best_rollno)
                    display_rollno = best_rollno
                    display_name = student['name']
                    display_status = f"Analyzing ({vote_count}/{VOTES_REQUIRED})"
                    display_color = (0, 255, 255)  # Yellow
            
            # Draw on all detected faces
            for (x, y, w, h) in faces:
                cv2.rectangle(frame, (x, y), (x+w, y+h), display_color, 2)
                
                if display_name:
                    # Show recognized or analyzing
                    name_text = f"{display_name} ({display_rollno})"
                    cv2.putText(frame, name_text, (x, y-30),
                               cv2.FONT_HERSHEY_SIMPLEX, 0.7, display_color, 2)
                    cv2.putText(frame, display_status, (x, y-10),
                               cv2.FONT_HERSHEY_SIMPLEX, 0.5, display_color, 2)
                    
                    # Show confidence and votes
                    if vote_count > 0:
                        info_text = f"Conf: {avg_confidence:.1f} | Votes: {vote_count}/{VOTES_REQUIRED}"
                        cv2.putText(frame, info_text, (x, y+h+20),
                                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
                else:
                    # Show unknown
                    cv2.putText(frame, "Unknown", (x, y-10),
                               cv2.FONT_HERSHEY_SIMPLEX, 0.7, display_color, 2)
                    
                    # Show current prediction for debugging
                    if current_confidence is not None:
                        debug_text = f"Conf: {current_confidence:.1f} (need <{CONFIDENCE_THRESHOLD})"
                        cv2.putText(frame, debug_text, (x, y+h+20),
                                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 0), 1)
            
            # Instructions overlay
            if recognized_rollno:
                status_text = f"Ready: {recognized_name} | SPACE: Confirm | ESC: Cancel"
                cv2.putText(frame, status_text, (10, 30),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
            else:
                cv2.putText(frame, "SPACE: Confirm | ESC: Cancel", (10, 30),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
            
            cv2.imshow('Take Attendance', frame)
            
            key = cv2.waitKey(1) & 0xFF
            if key == ord(' '):
                # User confirms
                if recognized_rollno:
                    break
                else:
                    print("⚠ No face recognized yet. Please look at camera...")
            elif key == 27:  # ESC
                cap.release()
                cv2.destroyAllWindows()
                return jsonify({'success': False, 'message': 'Attendance cancelled'})
        
        cap.release()
        cv2.destroyAllWindows()
        
        if not recognized_rollno:
            return jsonify({'success': False, 'message': 'Face not recognized. Please ensure good lighting and look directly at camera.'})
        
        print(f"\n✓ Recognized: {recognized_name} ({recognized_rollno})")
        print(f"  Confidence: {confidence_val:.1f}%")
        
        # Apply attendance logic
        today_recs = get_today_records()
        existing = None
        for r in today_recs:
            if r['rollno'] == recognized_rollno:
                existing = r
                break
        
        action, status = decide_attendance_action(existing)
        
        if action == 'CLOSED':
            return jsonify({'success': False, 'message': 'Attendance window is closed'})
        
        if action == 'IN':
            rec = save_or_update_attendance(
                rollno=recognized_rollno,
                status='PRESENT',
                intime=get_now_time_str()
            )
            return jsonify({
                'success': True,
                'message': f'✓ IN marked for {recognized_name} ({recognized_rollno}) at {rec["intime"]}',
                'name': recognized_name,
                'rollno': recognized_rollno,
                'confidence': confidence_val,
                'status': rec['status'],
                'intime': rec['intime'],
                'outtime': rec['outtime']
            })
        
        if action == 'OUT':
            rec = save_or_update_attendance(
                rollno=recognized_rollno,
                status=status,
                outtime=get_now_time_str()
            )
            return jsonify({
                'success': True,
                'message': f'✓ OUT marked for {recognized_name} ({recognized_rollno}) at {rec["outtime"]} ({status})',
                'name': recognized_name,
                'rollno': recognized_rollno,
                'confidence': confidence_val,
                'status': rec['status'],
                'intime': rec['intime'],
                'outtime': rec['outtime']
            })
            
    except Exception as e:
        print('Error in take_attendance:', e)
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Attendance error: {str(e)}'}), 500

@app.route('/api/todays_attendance')
def todays_attendance():
    today = get_today_date_str()
    students = load_students()
    all_att = load_attendance()
    
    today_map = {}
    for r in all_att:
        if r['date'] == today:
            today_map[r['rollno']] = r
    
    rows = []
    present_count = 0
    absent_count = 0
    
    for idx, s in enumerate(students, start=1):
        roll = s['rollno']
        rec = today_map.get(roll)
        
        if rec:
            status = rec['status']
            intime = rec['intime'] or '-'
            outtime = rec['outtime'] or '-'
        else:
            status = 'ABSENT'
            intime = '-'
            outtime = '-'
        
        if status == 'PRESENT' or status == 'HALF DAY':
            present_count += 1
        else:
            absent_count += 1
        
        rows.append({
            'sno': idx,
            'id': roll,
            'name': s['name'],
            'intime': intime,
            'outtime': outtime,
            'status': status
        })
    
    return jsonify({
        'success': True,
        'date': today,
        'total_students': len(students),
        'present': present_count,
        'absent': absent_count,
        'records': rows
    })
def create_attendance_excel(records, title, date_or_range):
    """Create Excel for daily attendance."""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    
    # Colors
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Date range
    ws.merge_cells('A2:F2')
    date_cell = ws['A2']
    date_cell.value = date_or_range
    date_cell.font = Font(italic=True)
    date_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['S.NO', 'ID', 'NAME', 'IN TIME', 'OUT TIME', 'STATUS']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_idx, rec in enumerate(records, 5):
        cells_data = [
            rec.get('sno', ''),
            rec.get('id', ''),
            rec.get('name', ''),
            rec.get('intime', '-'),
            rec.get('outtime', '-'),
            rec.get('status', '')
        ]
        
        for col, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col in [1, 6] else 'left')
            
            if col == 6:  # STATUS column
                if value == 'PRESENT':
                    cell.fill = green_fill
                elif value == 'HALF DAY':
                    cell.fill = yellow_fill
                elif value == 'ABSENT':
                    cell.fill = red_fill
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    return wb


def create_student_report_excel(student_name, student_id, start_date, end_date, records, summary_stats):
    """
    Create individual student attendance report Excel.
    summary_stats = {
        'total_days': int,
        'present': int,
        'half_days': int,
        'absent': int,
        'attendance_percentage': float,
        'is_defaulter': bool
    }
    """
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Report"
    
    # Colors
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    summary_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="2C3E50")
    summary_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== TITLE SECTION =====
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Student Report - {student_name} ({student_id})"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Date range
    ws.merge_cells('A2:F2')
    date_cell = ws['A2']
    date_cell.value = f"From {start_date} To {end_date}"
    date_cell.font = Font(italic=True, size=11)
    date_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18
    
    # ===== SUMMARY SECTION =====
    # Summary on left side (A-C)
    summary_row = 4
    summary_data = [
        ('Student', f'{student_name} (ID: {student_id})'),
        ('Date Range', f'{start_date} to {end_date}'),
        ('Total Days', str(summary_stats['total_days'])),
        ('Present', str(summary_stats['present'])),
        ('Half Days', str(summary_stats['half_days'])),
        ('Absent', str(summary_stats['absent'])),
        ('Attendance %', f"{summary_stats['attendance_percentage']}%"),
        ('Status', 'DEFAULTER' if summary_stats['is_defaulter'] else 'GOOD STANDING')
    ]
    
    for label, value in summary_data:
        # Label cell
        label_cell = ws.cell(row=summary_row, column=1)
        label_cell.value = label
        label_cell.font = Font(bold=True, size=10)
        label_cell.fill = summary_fill
        label_cell.border = border
        label_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Value cell
        value_cell = ws.cell(row=summary_row, column=2)
        value_cell.value = value
        value_cell.font = Font(size=10)
        value_cell.fill = summary_fill
        value_cell.border = border
        value_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        summary_row += 1
    
    # ===== ATTENDANCE TABLE SECTION =====
    table_start_row = 13
    
    # Table headers
    headers = ['S.NO', 'DATE', 'IN TIME', 'OUT TIME', 'STATUS', 'LATE?', 'EARLY?']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_start_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.row_dimensions[table_start_row].height = 20
    
    # Table data rows
    for row_idx, rec in enumerate(records, table_start_row + 1):
        cells_data = [
            rec.get('sno', ''),
            rec.get('date', ''),
            rec.get('intime', '-'),
            rec.get('outtime', '-'),
            rec.get('status', ''),
            'YES' if rec.get('islate', False) else 'NO',
            'YES' if rec.get('isearly', False) else 'NO'
        ]
        
        for col, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col in [1, 5, 6, 7] else 'left', vertical='center')
            
            # Color status column
            if col == 5:  # STATUS column
                if value == 'PRESENT':
                    cell.fill = green_fill
                elif value == 'HALF DAY':
                    cell.fill = yellow_fill
                elif value == 'ABSENT':
                    cell.fill = red_fill
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    
    return wb

@app.route('/api/attendance_dates')
def attendance_dates():
    all_att = load_attendance()
    dates = sorted({r['date'] for r in all_att}, reverse=True)
    return jsonify({'success': True, 'dates': dates})
@app.route('/api/attendancebydate')
def attendance_by_date():
    date = request.args.get('date')
    if not date:
        return jsonify({'success': False, 'message': 'Date is required'}), 400

    all_att = load_attendance()
    students = load_students()
    name_map = {s['rollno']: s['name'] for s in students}

    day_recs = [r for r in all_att if r['date'] == date]
    rows = []
    for i, r in enumerate(day_recs, start=1):
        rows.append({
            'sno': i,
            'id': r['rollno'],
            'name': name_map.get(r['rollno'], 'Unknown'),
            'intime': r['intime'] or '-',
            'outtime': r['outtime'] or '-',
            'status': r['status'],
            'date': r['date']
        })

    return jsonify({'success': True, 'date': date, 'records': rows})
@app.route('/api/exportattendance')
def export_attendance():
    date = request.args.get('date')
    if not date:
        return jsonify({'success': False, 'message': 'Date is required'}), 400

    all_att = load_attendance()
    students = load_students()
    name_map = {s['rollno']: s['name'] for s in students}

    day_recs = [r for r in all_att if r['date'] == date]
    records = []
    for i, r in enumerate(day_recs, start=1):
        records.append({
            'sno': i,
            'id': r['rollno'],
            'name': name_map.get(r['rollno'], 'Unknown'),
            'intime': r['intime'] or '-',
            'outtime': r['outtime'] or '-',
            'status': r['status'],
            'date': r['date']
        })

    wb = create_attendance_excel(records, f'Attendance Report - {date}', f'Date: {date}')
    
    filename = f'attendance_{date}.xlsx'
    filepath = f'exports/{filename}'
    os.makedirs('exports', exist_ok=True)
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name=filename)
@app.route('/api/students_analytics')
def students_analytics():
    students = load_students()
    all_att = load_attendance()

    stats = {}
    for s in students:
        stats[s['rollno']] = {
            'name': s['name'],
            'present': 0,
            'half': 0,
            'absent': 0,
            'total': 0
        }

    for r in all_att:
        roll = r['rollno']
        if roll not in stats:
            continue
        if r['status'] == 'PRESENT':
            stats[roll]['present'] += 1
            stats[roll]['total'] += 1
        elif r['status'] == 'HALF DAY':
            stats[roll]['half'] += 1
            stats[roll]['total'] += 1
        elif r['status'] == 'ABSENT':
            stats[roll]['absent'] += 1
            stats[roll]['total'] += 1

    defaulters = []
    for roll, st in stats.items():
        if st['total'] == 0:
            attendance_pct = 0.0
        else:
            effective_present = st['present'] + 0.5 * st['half']
            attendance_pct = round(100.0 * effective_present / st['total'], 2)

        if attendance_pct < 75.0:
            defaulters.append({
                'id': roll,
                'name': st['name'],
                'attendancepercentage': attendance_pct,
                'presentdays': st['present'],
                'absentdays': st['absent'],
                'halfdays': st['half']
            })

    return jsonify({'success': True, 'defaulters': defaulters})
@app.route('/api/individualreport/<student_id>')
def individual_report(student_id):
    start = request.args.get('startdate')
    end = request.args.get('enddate')
    if not start or not end:
        return jsonify({'error': 'Start and end dates are required'}), 400

    stu = find_student_by_roll(student_id)
    if not stu:
        return jsonify({'error': 'Student not found'}), 404

    all_att = load_attendance()
    history = []
    total_days = 0
    present = half = absent = 0

    for r in all_att:
        if r['rollno'] != student_id:
            continue
        if not (start <= r['date'] <= end):
            continue
        total_days += 1
        status = r['status']
        if status == 'PRESENT':
            present += 1
        elif status == 'HALF DAY':
            half += 1
        elif status == 'ABSENT':
            absent += 1

        history.append({
            'date': r['date'],
            'intime': r['intime'] or '-',
            'outtime': r['outtime'] or '-',
            'status': status,
            'islate': False,
            'isearly': False
        })

    if total_days == 0:
        attendance_pct = 0.0
    else:
        effective_present = present + 0.5 * half
        attendance_pct = round(100.0 * effective_present / total_days, 2)

    return jsonify({
        'studentid': student_id,
        'studentname': stu['name'],
        'totaldays': total_days,
        'presentdays': present,
        'halfdays': half,
        'absentdays': absent,
        'attendancepercentage': attendance_pct,
        'isdefaulter': attendance_pct < 75.0,
        'attendancehistory': history
    })
@app.route('/api/exportstudentreport/<student_id>')
def export_student_report(student_id):
    start = request.args.get('startdate')
    end = request.args.get('enddate')
    if not start or not end:
        return jsonify({'error': 'Dates required'}), 400

    stu = find_student_by_roll(student_id)
    if not stu:
        return jsonify({'error': 'Student not found'}), 404

    all_att = load_attendance()
    records = []
    total_days = 0
    present = half = absent = 0

    for r in all_att:
        if r['rollno'] != student_id or not (start <= r['date'] <= end):
            continue
        
        total_days += 1
        status = r['status']
        
        if status == 'PRESENT':
            present += 1
        elif status == 'HALF DAY':
            half += 1
        elif status == 'ABSENT':
            absent += 1

        records.append({
            'sno': len(records) + 1,
            'date': r['date'],
            'intime': r['intime'] or '-',
            'outtime': r['outtime'] or '-',
            'status': status,
            'islate': False,
            'isearly': False
        })

    if total_days == 0:
        attendance_pct = 0.0
    else:
        effective_present = present + 0.5 * half
        attendance_pct = round(100.0 * effective_present / total_days, 2)

    summary_stats = {
        'total_days': total_days,
        'present': present,
        'half_days': half,
        'absent': absent,
        'attendance_percentage': attendance_pct,
        'is_defaulter': attendance_pct < 75.0
    }

    wb = create_student_report_excel(
        student_name=stu['name'],
        student_id=student_id,
        start_date=start,
        end_date=end,
        records=records,
        summary_stats=summary_stats
    )

    filename = f"student_report_{student_id}_{start}_to_{end}.xlsx"
    filepath = f'exports/{filename}'
    os.makedirs('exports', exist_ok=True)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)


if __name__ == '__main__':
    app.run(debug=True, port=5000)
