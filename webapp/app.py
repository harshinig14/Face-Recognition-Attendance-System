from flask import Flask, render_template, send_file, jsonify, request
from flask_cors import CORS
import sys
from pathlib import Path
from datetime import datetime, timedelta
import json
import threading
import webbrowser
import time
from collections import defaultdict, Counter

# --- Paths ---
BASE_DIR = Path(__file__).parent.parent  # COE/attendance/
DATA_DIR = BASE_DIR / 'python-module' / 'data'
WEB_APP_DIR = Path(__file__).parent

# Ensure python-module/webapp is importable
sys.path.append(str(WEB_APP_DIR))

try:
    from simple_camera_service import SimpleCameraService
except Exception as e:
    print(f"Could not import SimpleCameraService: {e}")
    raise

app = Flask(__name__)
app.secret_key = 'your-secret-key-12345'
CORS(app)

# Import the improved camera service
camera_service = SimpleCameraService()

STUDENTS_FILE = DATA_DIR / 'students.json'
ATTENDANCE_FILE = DATA_DIR / 'attendance.json'
LOG_FILE = DATA_DIR / 'logs' / 'recognition_log.json'

def load_students():
    """Load students from JSON (resilient to formats)"""
    try:
        if STUDENTS_FILE.exists():
            with open(STUDENTS_FILE, 'r') as f:
                content = f.read().strip()
                if not content:
                    return []
                data = json.loads(content)
                if isinstance(data, list):
                    return data
                elif isinstance(data, dict):
                    return [data] if data else []
        return []
    except Exception as e:
        print(f"Error loading students.json: {e}")
        return []

def load_attendance_full():
    """Return the attendance data structure (dict)"""
    try:
        if ATTENDANCE_FILE.exists():
            with open(ATTENDANCE_FILE, 'r') as f:
                content = f.read().strip()
                if not content:
                    return {}
                return json.loads(content)
        return {}
    except Exception as e:
        print(f"Error reading attendance.json: {e}")
        return {}

@app.route('/')
def index():
    students = load_students()
    attendance = load_attendance_full()
    today = datetime.now().strftime('%Y-%m-%d')
    todays_attendance = attendance.get(today, {})
    
    # Convert dict to list for template
    attendance_list = []
    for student_id, record in todays_attendance.items():
        in_time = record.get('in_time') or record.get('intime') or '-'
        out_time = record.get('out_time') or record.get('outtime') or '-'
        attendance_list.append({
            'id': student_id,
            'name': record.get('name', 'Unknown'),
            'intime': in_time,
            'outtime': out_time,
            'status': record.get('status', 'ABSENT')
        })
    
    # Calculate counts
    present_count = sum(1 for r in todays_attendance.values()
                       if r.get('status') in ['PRESENT', 'HALF DAY'])
    absent_count = sum(1 for r in todays_attendance.values()
                      if r.get('status') == 'ABSENT')
    
    return render_template(
        'dashboard.html',
        teacher_username='Teacher',
        today=today,
        total_students=len(students),
        present_count=present_count,
        absent_count=absent_count,
        students_list=students,
        attendance_records=attendance_list
    )

@app.route('/dashboard')
def dashboard():
    return index()

@app.route('/api/start_camera')
def start_camera():
    success = camera_service.start_camera()
    return jsonify({'success': success})

@app.route('/api/stop_camera')
def stop_camera():
    camera_service.stop_camera()
    return jsonify({'success': True})

@app.route('/api/current_recognition')
def current_recognition():
    status = camera_service.get_status()
    return jsonify(status)

# ✅ FIXED: Get all students with detailed attendance analytics
@app.route('/api/students_analytics')
def students_analytics():
    """Calculate attendance analytics for all students including defaulters"""
    attendance = load_attendance_full()
    students = load_students()
    
    students_data = []
    defaulters = []
    
    # ✅ FIX: Get total working days (all dates in attendance system)
    total_working_days = len(attendance)
    
    for student in students:
        student_id = str(student.get('id'))
        student_name = student.get('name', 'Unknown')
        
        # Count attendance status for THIS student only
        present_days = 0
        half_days = 0
        absent_days = 0
        late_arrivals = 0
        early_departures = 0
        
        # ✅ FIX: Count student's individual attendance across all dates
        for date, day_records in attendance.items():
            if student_id in day_records:
                record = day_records[student_id]
                status = record.get('status', 'ABSENT')
                in_time = record.get('in_time')
                out_time = record.get('out_time')
                
                # Count status
                if status == 'PRESENT':
                    present_days += 1
                elif status == 'HALF DAY':
                    half_days += 1
                elif status == 'ABSENT':
                    absent_days += 1
                
                # Check for late arrival (after 8:00 AM)
                if in_time and in_time != '-':
                    try:
                        hour = int(in_time.split(':')[0])
                        minute = int(in_time.split(':')[1])
                        if hour > 8 or (hour == 8 and minute > 0):
                            late_arrivals += 1
                    except:
                        pass
                
                # Check for early departure (before 2:50 PM)
                if out_time and out_time != '-':
                    try:
                        hour = int(out_time.split(':')[0])
                        minute = int(out_time.split(':')[1])
                        if hour < 14 or (hour == 14 and minute < 50):
                            early_departures += 1
                    except:
                        pass
            else:
                # Student not marked for this day = absent
                absent_days += 1
        
        # ✅ FIX: Calculate attendance percentage correctly
        # Formula: (Present + 0.5*HalfDay) / TotalWorkingDays * 100
        attended_days = present_days + (half_days * 0.5)
        attendance_percentage = (attended_days / total_working_days * 100) if total_working_days > 0 else 0
        
        student_data = {
            'id': student_id,
            'name': student_name,
            'total_days': total_working_days,
            'present_days': present_days,
            'half_days': half_days,
            'absent_days': absent_days,
            'late_arrivals': late_arrivals,
            'early_departures': early_departures,
            'attendance_percentage': round(attendance_percentage, 2),
            'is_defaulter': attendance_percentage < 75.0
        }
        
        students_data.append(student_data)
        
        # Add to defaulters list if below threshold
        if attendance_percentage < 75.0:
            defaulters.append(student_data)
    
    return jsonify({
        'students': students_data,
        'defaulters': defaulters,
        'total_students': len(students),
        'total_defaulters': len(defaulters),
        'threshold': 75.0
    })

# ✅ FIXED: Get detailed individual student report with PROPER date range filter
@app.route('/api/individual_report/<student_id>')
def individual_report(student_id):
    """Generate detailed report for individual student with proper date filtering"""
    attendance = load_attendance_full()
    students = load_students()
    
    # Get date range parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    student = next((s for s in students if str(s['id']) == str(student_id)), None)
    if not student:
        return jsonify({'error': 'Student not found'}), 404
    
    student_name = student.get('name', 'Unknown')
    
    # Detailed attendance history - ONLY within date range
    attendance_history = []
    present_days = 0
    half_days = 0
    absent_days = 0
    late_arrivals = 0
    early_departures = 0
    
    # Get all dates in the attendance system that are within the specified range
    all_dates_in_range = []
    
    for date in sorted(attendance.keys()):
        # Skip dates outside the specified range
        if start_date and date < start_date:
            continue
        if end_date and date > end_date:
            continue
        all_dates_in_range.append(date)
    
    # Now process only the dates in the range
    for date in all_dates_in_range:
        day_records = attendance[date]
        
        if student_id in day_records:
            record = day_records[student_id]
            status = record.get('status', 'ABSENT')
            in_time = record.get('in_time', '-')
            out_time = record.get('out_time', '-')
            
            # Determine if late/early
            is_late = False
            is_early = False
            
            if in_time and in_time != '-':
                try:
                    hour = int(in_time.split(':')[0])
                    minute = int(in_time.split(':')[1])
                    if hour > 8 or (hour == 8 and minute > 0):
                        is_late = True
                        late_arrivals += 1
                except:
                    pass
            
            if out_time and out_time != '-':
                try:
                    hour = int(out_time.split(':')[0])
                    minute = int(out_time.split(':')[1])
                    if hour < 14 or (hour == 14 and minute < 50):
                        is_early = True
                        early_departures += 1
                except:
                    pass
            
            attendance_history.append({
                'date': date,
                'in_time': in_time,
                'out_time': out_time,
                'status': status,
                'is_late': is_late,
                'is_early': is_early
            })
            
            if status == 'PRESENT':
                present_days += 1
            elif status == 'HALF DAY':
                half_days += 1
            elif status == 'ABSENT':
                absent_days += 1
        else:
            # Student not marked for this day = absent
            attendance_history.append({
                'date': date,
                'in_time': '-',
                'out_time': '-',
                'status': 'ABSENT',
                'is_late': False,
                'is_early': False
            })
            absent_days += 1
    
    # Calculate percentage based ONLY on dates in the specified range
    total_days_in_range = len(all_dates_in_range)
    
    if total_days_in_range == 0:
        # If no dates in range, try to get some default range
        if not start_date and not end_date:
            # If no date range specified, use all dates
            total_days_in_range = len(attendance)
        else:
            total_days_in_range = 0
    
    attended_days = present_days + (half_days * 0.5)
    attendance_percentage = (attended_days / total_days_in_range * 100) if total_days_in_range > 0 else 0
    
    return jsonify({
        'student_id': student_id,
        'student_name': student_name,
        'total_days': total_days_in_range,
        'present_days': present_days,
        'half_days': half_days,
        'absent_days': absent_days,
        'late_arrivals': late_arrivals,
        'early_departures': early_departures,
        'attendance_percentage': round(attendance_percentage, 2),
        'is_defaulter': attendance_percentage < 75.0,
        'attendance_history': attendance_history,
        'date_range': {
            'start_date': start_date,
            'end_date': end_date
        }
    })

# ✅ UPDATED: Export individual student report to Excel with PROPER date range filter
@app.route('/api/export_student_report/<student_id>')
def export_student_report(student_id):
    """Export individual student report to Excel with proper date range filtering"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    try:
        # Get date range parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Get student report data with PROPER date range filter
        report_response = individual_report(student_id)
        report_data = report_response.get_json()
        
        if 'error' in report_data:
            return jsonify({'error': 'Student not found'}), 404
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Report"
        
        # Title with date range info
        date_range_text = ""
        if start_date and end_date:
            date_range_text = f" (From {start_date} to {end_date})"
        elif start_date:
            date_range_text = f" (From {start_date})"
        elif end_date:
            date_range_text = f" (Until {end_date})"
        
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"INDIVIDUAL ATTENDANCE REPORT{date_range_text}"
        title_cell.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Student Info
        ws['A3'] = "Student ID:"
        ws['B3'] = report_data['student_id']
        ws['A4'] = "Student Name:"
        ws['B4'] = report_data['student_name']
        
        # Date Range Info
        if start_date or end_date:
            ws['A5'] = "Date Range:"
            range_text = ""
            if start_date and end_date:
                range_text = f"{start_date} to {end_date}"
            elif start_date:
                range_text = f"From {start_date}"
            elif end_date:
                range_text = f"Until {end_date}"
            ws['B5'] = range_text
        
        for cell in ['A3', 'A4', 'A5']:
            if cell in ['A3', 'A4'] or (cell == 'A5' and (start_date or end_date)):
                ws[cell].font = Font(name='Arial', size=11, bold=True)
        
        # Summary Statistics
        summary_start_row = 7 if (start_date or end_date) else 6
        ws[f'A{summary_start_row}'] = "ATTENDANCE SUMMARY"
        ws[f'A{summary_start_row}'].font = Font(name='Arial', size=12, bold=True)
        ws[f'A{summary_start_row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        summary_data = [
            ("Total Working Days:", report_data['total_days']),
            ("Present Days:", report_data['present_days']),
            ("Half Days:", report_data['half_days']),
            ("Absent Days:", report_data['absent_days']),
            ("Late Arrivals:", report_data['late_arrivals']),
            ("Early Departures:", report_data['early_departures']),
            ("Attendance %:", f"{report_data['attendance_percentage']}%"),
            ("Status:", "DEFAULTER (Below 75%)" if report_data['is_defaulter'] else "GOOD STANDING")
        ]
        
        for i, (label, value) in enumerate(summary_data, summary_start_row + 1):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(name='Arial', size=11, bold=True)
            
            # Highlight defaulter status
            if label == "Status:":
                if report_data['is_defaulter']:
                    ws[f'B{i}'].font = Font(name='Arial', size=11, bold=True, color="9C0006")
                    ws[f'B{i}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    ws[f'B{i}'].font = Font(name='Arial', size=11, bold=True, color="006100")
                    ws[f'B{i}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Attendance History Table
        header_row = summary_start_row + len(summary_data) + 2
        ws[f'A{header_row}'] = "DETAILED ATTENDANCE HISTORY"
        ws[f'A{header_row}'].font = Font(name='Arial', size=12, bold=True)
        
        headers = ["Date", "IN Time", "OUT Time", "Status", "Late?", "Early?"]
        header_row = header_row + 1
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows - ONLY the filtered data
        for idx, record in enumerate(report_data['attendance_history'], header_row + 1):
            ws.cell(row=idx, column=1, value=record['date'])
            ws.cell(row=idx, column=2, value=record['in_time'])
            ws.cell(row=idx, column=3, value=record['out_time'])
            
            status_cell = ws.cell(row=idx, column=4, value=record['status'])
            status_cell.font = Font(name='Arial', size=10, bold=True)
            
            if record['status'] == 'PRESENT':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(name='Arial', size=10, bold=True, color="006100")
            elif record['status'] == 'HALF DAY':
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                status_cell.font = Font(name='Arial', size=10, bold=True, color="9C6500")
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(name='Arial', size=10, bold=True, color="9C0006")
            
            ws.cell(row=idx, column=5, value="YES" if record['is_late'] else "NO")
            ws.cell(row=idx, column=6, value="YES" if record['is_early'] else "NO")
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        
        # Save file
        date_suffix = f"_{start_date}_to_{end_date}" if start_date and end_date else ""
        export_path = DATA_DIR / f'student_report_{student_id}{date_suffix}.xlsx'
        wb.save(export_path)
        
        # Create download filename
        download_name = f'student_report_{report_data["student_name"]}_{student_id}'
        if start_date and end_date:
            download_name += f'_{start_date}_to_{end_date}'
        download_name += '.xlsx'
        
        return send_file(
            export_path,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        print(f"Export error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/students_list')
def students_list():
    students = load_students()
    return jsonify({'students': students, 'total': len(students)})

@app.route('/api/todays_attendance')
def todays_attendance():
    attendance = load_attendance_full()
    today = datetime.now().strftime('%Y-%m-%d')
    todays_records = attendance.get(today, {})
    
    attendance_list = []
    for student_id, record in todays_records.items():
        in_time = record.get('in_time') or record.get('intime') or '-'
        out_time = record.get('out_time') or record.get('outtime') or '-'
        attendance_list.append({
            'id': student_id,
            'name': record.get('name', 'Unknown'),
            'intime': in_time,
            'outtime': out_time,
            'status': record.get('status', 'ABSENT')
        })
    
    present_count = sum(1 for r in todays_records.values()
                       if r.get('status') in ['PRESENT', 'HALF DAY'])
    absent_count = sum(1 for r in todays_records.values()
                      if r.get('status') == 'ABSENT')
    
    return jsonify({
        'date': today,
        'attendance': attendance_list,
        'present_count': present_count,
        'absent_count': absent_count
    })

@app.route('/api/attendance_by_date')
def attendance_by_date():
    date = request.args.get('date')
    if not date:
        return jsonify({'error': 'Date parameter required'}), 400
    
    attendance = load_attendance_full()
    day_records = attendance.get(date, {})
    
    attendance_list = []
    for student_id, record in day_records.items():
        in_time = record.get('in_time') or record.get('intime') or '-'
        out_time = record.get('out_time') or record.get('outtime') or '-'
        attendance_list.append({
            'id': student_id,
            'name': record.get('name', 'Unknown'),
            'intime': in_time,
            'outtime': out_time,
            'status': record.get('status', 'ABSENT')
        })
    
    present_count = sum(1 for r in day_records.values()
                       if r.get('status') in ['PRESENT', 'HALF DAY'])
    absent_count = sum(1 for r in day_records.values()
                      if r.get('status') == 'ABSENT')
    
    return jsonify({
        'date': date,
        'attendance': attendance_list,
        'present_count': present_count,
        'absent_count': absent_count
    })

@app.route('/api/attendance_dates')
def attendance_dates():
    attendance = load_attendance_full()
    dates = sorted(attendance.keys(), reverse=True)
    return jsonify({'dates': dates})

@app.route('/api/weekly_report')
def weekly_report():
    attendance = load_attendance_full()
    students = load_students()
    total_students = len(students)
    
    today = datetime.now()
    week_data = []
    
    for i in range(6, -1, -1):
        date = (today - timedelta(days=i)).strftime('%Y-%m-%d')
        day_records = attendance.get(date, {})
        present = sum(1 for r in day_records.values() if r.get('status') in ['PRESENT', 'HALF DAY'])
        absent = sum(1 for r in day_records.values() if r.get('status') == 'ABSENT')
        
        week_data.append({
            'date': date,
            'present': present,
            'absent': absent,
            'total': total_students
        })
    
    return jsonify({'weekly_data': week_data})

@app.route('/api/monthly_report')
def monthly_report():
    attendance = load_attendance_full()
    students = load_students()
    total_students = len(students)
    
    today = datetime.now()
    month_start = today.replace(day=1)
    days_in_month = (month_start.replace(month=month_start.month % 12 + 1, day=1) - timedelta(days=1)).day if month_start.month < 12 else 31
    
    month_data = []
    status_summary = Counter()
    
    for day in range(1, min(today.day + 1, days_in_month + 1)):
        date = today.replace(day=day).strftime('%Y-%m-%d')
        day_records = attendance.get(date, {})
        
        present = sum(1 for r in day_records.values() if r.get('status') == 'PRESENT')
        half_day = sum(1 for r in day_records.values() if r.get('status') == 'HALF DAY')
        absent = sum(1 for r in day_records.values() if r.get('status') == 'ABSENT')
        
        status_summary['PRESENT'] += present
        status_summary['HALF DAY'] += half_day
        status_summary['ABSENT'] += absent
        
        month_data.append({
            'date': date,
            'present': present,
            'half_day': half_day,
            'absent': absent,
            'total': total_students
        })
    
    return jsonify({
        'monthly_data': month_data,
        'summary': dict(status_summary)
    })

@app.route('/api/student_report/<student_id>')
def student_report(student_id):
    attendance = load_attendance_full()
    students = load_students()
    
    student_name = next((s['name'] for s in students if str(s['id']) == str(student_id)), 'Unknown')
    
    student_attendance = []
    status_count = Counter()
    
    for date, day_records in sorted(attendance.items()):
        if student_id in day_records:
            record = day_records[student_id]
            status = record.get('status', 'ABSENT')
            status_count[status] += 1
            
            student_attendance.append({
                'date': date,
                'in_time': record.get('in_time', '-'),
                'out_time': record.get('out_time', '-'),
                'status': status
            })
    
    total_days = len(attendance)
    present_days = status_count['PRESENT'] + status_count.get('HALF DAY', 0)
    attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0
    
    return jsonify({
        'student_id': student_id,
        'student_name': student_name,
        'attendance_history': student_attendance,
        'summary': dict(status_count),
        'attendance_percentage': round(attendance_percentage, 2),
        'total_days': total_days
    })

# ✅ NEW: Enhanced Excel export with improved formatting
@app.route('/api/export_attendance')
def export_attendance():
    """Export attendance to Excel with improved formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        attendance = load_attendance_full()
        
        if date not in attendance:
            return jsonify({'error': 'No attendance data for this date'}), 404
        
        day_records = attendance[date]
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"ATTENDANCE REPORT - {date}"
        title_cell.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Headers
        headers = ["S.No", "Roll No", "Name", "IN Time", "OUT Time", "Status"]
        header_row = 2
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        ws.row_dimensions[header_row].height = 25
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Data rows
        records = []
        for sid, rec in day_records.items():
            records.append({
                'id': sid,
                'name': rec.get('name', 'Unknown'),
                'in_time': rec.get('in_time', '-'),
                'out_time': rec.get('out_time', '-'),
                'status': rec.get('status', 'ABSENT')
            })
        
        # Sort by roll number
        records.sort(key=lambda x: int(x['id']) if x['id'].isdigit() else 0)
        
        for idx, record in enumerate(records, 1):
            row_num = idx + 2
            
            # S.No
            cell = ws.cell(row=row_num, column=1, value=idx)
            cell.alignment = Alignment(horizontal='center')
            
            # Roll No
            cell = ws.cell(row=row_num, column=2, value=record['id'])
            cell.alignment = Alignment(horizontal='center')
            
            # Name
            cell = ws.cell(row=row_num, column=3, value=record['name'])
            cell.alignment = Alignment(horizontal='left')
            
            # IN Time
            cell = ws.cell(row=row_num, column=4, value=record['in_time'])
            cell.alignment = Alignment(horizontal='center')
            
            # OUT Time
            cell = ws.cell(row=row_num, column=5, value=record['out_time'])
            cell.alignment = Alignment(horizontal='center')
            
            # Status with color coding
            status = record['status']
            cell = ws.cell(row=row_num, column=6, value=status)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name='Arial', size=11, bold=True)
            
            if status == 'PRESENT':
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(name='Arial', size=11, bold=True, color="006100")
            elif status == 'HALF DAY':
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                cell.font = Font(name='Arial', size=11, bold=True, color="9C6500")
            elif status == 'ABSENT':
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(name='Arial', size=11, bold=True, color="9C0006")
            
            # Apply borders to all cells in row
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Summary section
        summary_row = len(records) + 4
        ws.merge_cells(f'A{summary_row}:B{summary_row}')
        summary_cell = ws[f'A{summary_row}']
        summary_cell.value = "SUMMARY"
        summary_cell.font = Font(name='Arial', size=12, bold=True)
        summary_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Count statistics
        present_count = sum(1 for r in records if r['status'] == 'PRESENT')
        half_day_count = sum(1 for r in records if r['status'] == 'HALF DAY')
        absent_count = sum(1 for r in records if r['status'] == 'ABSENT')
        total_students = len(records)
        
        summary_data = [
            ("Total Students:", total_students),
            ("Present:", present_count),
            ("Half Day:", half_day_count),
            ("Absent:", absent_count),
            ("Attendance %:", f"{(present_count + half_day_count) / total_students * 100:.1f}%" if total_students > 0 else "0%")
        ]
        
        for i, (label, value) in enumerate(summary_data):
            row = summary_row + i + 1
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
        
        # Save file
        export_path = DATA_DIR / f'attendance_{date}.xlsx'
        wb.save(export_path)
        
        return send_file(
            export_path,
            as_attachment=True,
            download_name=f'attendance_{date}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        print(f"Export error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/recognition_logs')
def recognition_logs():
    """Return recognition logs JSON"""
    try:
        if LOG_FILE.exists():
            with open(LOG_FILE, 'r') as f:
                content = f.read().strip()
                if not content:
                    return jsonify([])
                data = json.loads(content)
                return jsonify(data)
        else:
            return jsonify([])
    except Exception as e:
        print(f"Error reading log file: {e}")
        return jsonify([]), 500


@app.route('/api/status')
def status():
    """Status endpoint for dashboard"""
    try:
        status = camera_service.get_status()
        students = load_students()
        attendance = load_attendance_full()
        today = datetime.now().strftime('%Y-%m-%d')
        todays_records = attendance.get(today, {})
        
        present_count = sum(1 for r in todays_records.values()
                           if r.get('status') in ['PRESENT', 'HALF DAY'])
        absent_count = sum(1 for r in todays_records.values()
                          if r.get('status') == 'ABSENT')
        
        status.update({
            'total_students': len(students),
            'present_today': present_count,
            'absent_today': absent_count
        })
        return jsonify(status)
    except Exception as e:
        print(f"Status error: {e}")
        return jsonify({}), 500

if __name__ == '__main__':
    def open_browser():
        time.sleep(1)
        try:
            webbrowser.open('http://localhost:5000')
        except:
            pass
    
    threading.Thread(target=open_browser, daemon=True).start()
    
    print("=" * 60)
    print("🎓 FACE RECOGNITION ATTENDANCE - STARTING")
    print(f"📁 DATA_DIR: {DATA_DIR}")
    print(f"📋 Students file: {STUDENTS_FILE.exists()}")
    print(f"📊 Attendance file: {ATTENDANCE_FILE.exists()}")
    print("=" * 60)
    
    app.run(host='0.0.0.0', port=5000, debug=True, use_reloader=False)